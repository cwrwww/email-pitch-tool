"""Email Pitch Tool - 轻量级邮件营销工具 MVP"""
import os
import io
import csv
import json
import sqlite3
import base64
from datetime import datetime, timedelta
from pathlib import Path
from contextlib import contextmanager
from typing import Optional
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from jinja2 import Template
from openpyxl import load_workbook
from email_validator import validate_email, EmailNotValidError
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from apscheduler.schedulers.background import BackgroundScheduler
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from googleapiclient.discovery import build

app = FastAPI(title="Email Pitch Tool")
scheduler = BackgroundScheduler()
scheduler.start()

# 配置
DB_PATH = "data.db"
CREDENTIALS_FILE = "credentials.json"  # 从Google Cloud Console下载
SCOPES = ['https://www.googleapis.com/auth/gmail.send', 'https://www.googleapis.com/auth/gmail.readonly']
# 支持环境变量配置REDIRECT_URI（部署时使用）
BASE_URL = os.environ.get("BASE_URL", "http://localhost:8000")
REDIRECT_URI = f"{BASE_URL}/oauth/callback"
TEST_MODE = os.environ.get("TEST_MODE", "false").lower() == "true"  # 测试模式不发真邮件

# 数据库初始化
def init_db():
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    # 启用WAL模式以提高并发性能
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")  # 30秒超时
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS accounts (
            id INTEGER PRIMARY KEY, email TEXT UNIQUE, token TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS campaigns (
            id INTEGER PRIMARY KEY, name TEXT, status TEXT DEFAULT 'draft',
            account_email TEXT, interval_minutes INTEGER DEFAULT 5,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS templates (
            id INTEGER PRIMARY KEY, campaign_id INTEGER, step INTEGER, subject TEXT, body TEXT,
            delay_days INTEGER DEFAULT 0, FOREIGN KEY(campaign_id) REFERENCES campaigns(id)
        );
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY, campaign_id INTEGER, email TEXT, data TEXT,
            status TEXT DEFAULT 'pending', current_step INTEGER DEFAULT 1,
            last_sent_at TIMESTAMP, opened INTEGER DEFAULT 0, clicked INTEGER DEFAULT 0, replied INTEGER DEFAULT 0,
            FOREIGN KEY(campaign_id) REFERENCES campaigns(id)
        );
        CREATE TABLE IF NOT EXISTS blacklist (email TEXT PRIMARY KEY);
    """)
    conn.commit()
    conn.close()
init_db()

def check_all_replies():
    """定期检查所有campaign的回复"""
    try:
        with get_db() as conn:
            campaigns = conn.execute(
                "SELECT DISTINCT c.id, c.account_email FROM campaigns c "
                "JOIN leads l ON c.id = l.campaign_id "
                "WHERE c.account_email IS NOT NULL AND l.replied = 0"
            ).fetchall()

        # 释放数据库连接后再逐个检查
        for campaign in campaigns:
            try:
                check_replies(campaign['id'], campaign['account_email'])
            except Exception as e:
                print(f"[Check replies error for campaign {campaign['id']}] {e}")
    except Exception as e:
        print(f"[Check all replies error] {e}")

def restore_running_campaigns():
    """启动时恢复所有运行中的campaigns"""
    with get_db() as conn:
        rows = conn.execute("SELECT id, account_email, interval_minutes FROM campaigns WHERE status='running'").fetchall()

    for row in rows:
        if row['account_email']:
            scheduler.add_job(process_campaign, 'interval', minutes=row['interval_minutes'] or 5,
                              args=[row['id'], row['account_email']], id=f"campaign_{row['id']}", replace_existing=True)
            print(f"[Restored] Campaign {row['id']} with interval {row['interval_minutes']}min")

    # 添加定期回复检查任务（每10分钟检查一次所有campaign）
    scheduler.add_job(check_all_replies, 'interval', minutes=10, id='check_all_replies', replace_existing=True)
    print("[Scheduled] Reply checker every 10 minutes")

@contextmanager
def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=30.0)
    conn.execute("PRAGMA busy_timeout=30000")  # 30秒超时
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

# Gmail OAuth
@app.get("/oauth/start")
def oauth_start():
    # 支持环境变量配置（用于云端部署）
    if os.environ.get("GOOGLE_CLIENT_ID") and os.environ.get("GOOGLE_CLIENT_SECRET"):
        client_config = {
            "web": {
                "client_id": os.environ.get("GOOGLE_CLIENT_ID"),
                "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET"),
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [REDIRECT_URI]
            }
        }
        flow = Flow.from_client_config(client_config, scopes=SCOPES, redirect_uri=REDIRECT_URI)
    elif Path(CREDENTIALS_FILE).exists():
        flow = Flow.from_client_secrets_file(CREDENTIALS_FILE, scopes=SCOPES, redirect_uri=REDIRECT_URI)
    else:
        return {"error": "请先配置 Google OAuth 凭据（credentials.json 或环境变量）"}

    auth_url, _ = flow.authorization_url(prompt='consent')
    return RedirectResponse(auth_url)

@app.get("/oauth/callback")
def oauth_callback(code: str):
    # 使用动态REDIRECT_URI
    redirect_uri = f"{BASE_URL}/oauth/callback"

    # 支持环境变量配置
    if os.environ.get("GOOGLE_CLIENT_ID") and os.environ.get("GOOGLE_CLIENT_SECRET"):
        client_config = {
            "web": {
                "client_id": os.environ.get("GOOGLE_CLIENT_ID"),
                "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET"),
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [redirect_uri]
            }
        }
        flow = Flow.from_client_config(client_config, scopes=SCOPES, redirect_uri=redirect_uri)
    else:
        flow = Flow.from_client_secrets_file(CREDENTIALS_FILE, scopes=SCOPES, redirect_uri=redirect_uri)

    flow.fetch_token(code=code)
    creds = flow.credentials
    service = build('gmail', 'v1', credentials=creds)
    profile = service.users().getProfile(userId='me').execute()
    email = profile['emailAddress']
    with get_db() as conn:
        conn.execute("INSERT OR REPLACE INTO accounts(email, token) VALUES(?, ?)",
                    (email, creds.to_json()))
        conn.commit()
    return RedirectResponse("/?msg=账号绑定成功: " + email)

# API 路由
@app.post("/api/campaigns")
def create_campaign(name: str = Form(...)):
    with get_db() as conn:
        cur = conn.execute("INSERT INTO campaigns(name) VALUES(?)", (name,))
        conn.commit()
        return {"id": cur.lastrowid, "name": name}

@app.get("/api/campaigns")
def list_campaigns():
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM campaigns ORDER BY id DESC").fetchall()
        return [dict(r) for r in rows]

@app.post("/api/campaigns/{cid}/templates")
def add_template(cid: int, step: int = Form(...), subject: str = Form(...), body: str = Form(...), delay_days: int = Form(0)):
    with get_db() as conn:
        conn.execute("INSERT INTO templates(campaign_id, step, subject, body, delay_days) VALUES(?,?,?,?,?)",
                    (cid, step, subject, body, delay_days))
        conn.commit()
    return {"ok": True}

@app.get("/api/campaigns/{cid}/templates")
def get_templates(cid: int):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM templates WHERE campaign_id=? ORDER BY step", (cid,)).fetchall()
        return [dict(r) for r in rows]

@app.put("/api/templates/{tid}")
def update_template(tid: int, step: int = Form(...), subject: str = Form(...), body: str = Form(...), delay_days: int = Form(0)):
    with get_db() as conn:
        conn.execute("UPDATE templates SET step=?, subject=?, body=?, delay_days=? WHERE id=?",
                    (step, subject, body, delay_days, tid))
        conn.commit()
    return {"ok": True}

@app.delete("/api/templates/{tid}")
def delete_template(tid: int):
    with get_db() as conn:
        conn.execute("DELETE FROM templates WHERE id=?", (tid,))
        conn.commit()
    return {"ok": True}

@app.get("/api/campaigns/{cid}/variables")
def get_campaign_variables(cid: int):
    """提取campaign所有模板中使用的变量"""
    import re
    variables = set()
    with get_db() as conn:
        rows = conn.execute("SELECT subject, body FROM templates WHERE campaign_id=?", (cid,)).fetchall()
        for row in rows:
            # 匹配 {{var}} 或 {{var|default}}
            for text in [row['subject'], row['body']]:
                if text:
                    matches = re.findall(r'\{\{(\w+)(?:\|[^}]*)?\}\}', text)
                    variables.update(matches)
    # email是必须的，不需要提示
    variables.discard('email')
    return {"variables": sorted(variables)}

@app.post("/api/campaigns/{cid}/leads")
async def upload_leads(cid: int, file: UploadFile = File(...), defaults: str = Form("{}")):
    content = await file.read()
    default_values = json.loads(defaults)
    rows = []

    if file.filename.endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    else:
        wb = load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

    if not rows or 'email' not in rows[0]:
        raise HTTPException(400, "文件必须包含email列")

    added, skipped = 0, 0
    with get_db() as conn:
        blacklist = {r[0] for r in conn.execute("SELECT email FROM blacklist").fetchall()}
        existing = {r[0] for r in conn.execute("SELECT email FROM leads WHERE campaign_id=?", (cid,)).fetchall()}

        for row in rows:
            email = str(row.get('email', '')).strip().lower()
            try:
                validate_email(email)
                if email in blacklist or email in existing:
                    skipped += 1; continue
                # 合并默认值和行数据
                data = {**default_values, **{k: v for k, v in row.items() if k != 'email'}}
                conn.execute("INSERT INTO leads(campaign_id, email, data) VALUES(?,?,?)",
                            (cid, email, json.dumps(data, default=str)))
                added += 1
            except EmailNotValidError:
                skipped += 1
        conn.commit()
    return {"added": added, "skipped": skipped}

@app.get("/api/campaigns/{cid}/leads")
def get_leads(cid: int):
    with get_db() as conn:
        rows = conn.execute("SELECT * FROM leads WHERE campaign_id=?", (cid,)).fetchall()
        return [dict(r) for r in rows]

@app.post("/api/campaigns/{cid}/leads/json")
async def add_leads_json(cid: int, request: Request):
    """支持JSON格式批量导入leads"""
    rows = await request.json()

    added, skipped = 0, 0
    errors = []
    with get_db() as conn:
        blacklist = {r[0] for r in conn.execute("SELECT email FROM blacklist").fetchall()}
        existing = {r[0] for r in conn.execute("SELECT email FROM leads WHERE campaign_id=?", (cid,)).fetchall()}

        for row in rows:
            email = str(row.get('email', '')).strip().lower()
            if not email:
                errors.append("空邮箱")
                skipped += 1
                continue
            try:
                # 使用宽松模式验证邮箱
                validate_email(email, check_deliverability=False)
                if email in blacklist:
                    errors.append(f"{email}: 在黑名单中")
                    skipped += 1
                    continue
                if email in existing:
                    errors.append(f"{email}: 已存在")
                    skipped += 1
                    continue
                data = {k: v for k, v in row.items() if k != 'email'}
                conn.execute("INSERT INTO leads(campaign_id, email, data) VALUES(?,?,?)",
                            (cid, email, json.dumps(data, default=str)))
                existing.add(email)
                added += 1
            except EmailNotValidError as e:
                errors.append(f"{email}: {str(e)}")
                skipped += 1
        conn.commit()
    return {"added": added, "skipped": skipped, "errors": errors[:5]}

@app.get("/api/campaigns/{cid}/preview")
def preview_email(cid: int, lead_id: int, step: int = 1):
    with get_db() as conn:
        lead = conn.execute("SELECT * FROM leads WHERE id=?", (lead_id,)).fetchone()
        tpl = conn.execute("SELECT * FROM templates WHERE campaign_id=? AND step=?", (cid, step)).fetchone()
        if not lead or not tpl: raise HTTPException(404)
        data = json.loads(lead['data'])
        data['email'] = lead['email']
        return {
            "subject": Template(tpl['subject']).render(**data),
            "body": Template(tpl['body']).render(**data)
        }

@app.post("/api/campaigns/{cid}/launch")
def launch_campaign(cid: int, account_email: str = Form(...), interval_minutes: int = Form(5)):
    with get_db() as conn:
        conn.execute("UPDATE campaigns SET status='running', account_email=?, interval_minutes=? WHERE id=?",
                    (account_email, interval_minutes, cid))
        conn.commit()
    scheduler.add_job(process_campaign, 'interval', minutes=interval_minutes,
                      args=[cid, account_email], id=f"campaign_{cid}", replace_existing=True)
    return {"ok": True, "msg": f"已启动，每{interval_minutes}分钟发送一封"}

@app.get("/api/campaigns/{cid}")
def get_campaign(cid: int):
    with get_db() as conn:
        row = conn.execute("SELECT * FROM campaigns WHERE id=?", (cid,)).fetchone()
        if not row: raise HTTPException(404)
        return dict(row)

@app.post("/api/campaigns/{cid}/stop")
def stop_campaign(cid: int):
    with get_db() as conn:
        conn.execute("UPDATE campaigns SET status='paused' WHERE id=?", (cid,))
        conn.commit()
    try: scheduler.remove_job(f"campaign_{cid}")
    except: pass
    return {"ok": True}

def send_gmail(account_email: str, to: str, subject: str, body: str) -> bool:
    if TEST_MODE:
        print(f"[TEST MODE] Would send email:")
        print(f"  From: {account_email}")
        print(f"  To: {to}")
        print(f"  Subject: {subject}")
        print(f"  Body: {body[:200]}...")
        return True

    with get_db() as conn:
        row = conn.execute("SELECT token FROM accounts WHERE email=?", (account_email,)).fetchone()
        if not row: return False
    creds = Credentials.from_authorized_user_info(json.loads(row['token']))
    service = build('gmail', 'v1', credentials=creds)
    msg = MIMEMultipart('alternative')
    msg['To'], msg['Subject'] = to, subject
    msg.attach(MIMEText(body, 'html'))
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId='me', body={'raw': raw}).execute()
    return True

def check_replies(cid: int, account_email: str):
    """检查收件箱，标记已回复的leads"""
    if TEST_MODE:
        print(f"[TEST MODE] Skipping reply check for campaign {cid}")
        return

    try:
        # 首先获取认证信息和leads列表（快速查询，避免长时间持有连接）
        with get_db() as conn:
            row = conn.execute("SELECT token FROM accounts WHERE email=?", (account_email,)).fetchone()
            if not row:
                return

            token_json = row['token']

            # 获取该campaign所有未回复的lead邮箱
            # 重要：只检查已经发送过邮件的leads (last_sent_at IS NOT NULL)
            leads = conn.execute(
                "SELECT id, email, last_sent_at FROM leads WHERE campaign_id=? AND replied=0 AND last_sent_at IS NOT NULL", (cid,)
            ).fetchall()
            if not leads:
                return

        # 释放数据库连接后再进行Gmail API调用（耗时操作）
        # 存储 {email: (lead_id, last_sent_at)}
        lead_emails = {l['email'].lower(): (l['id'], l['last_sent_at']) for l in leads}

        creds = Credentials.from_authorized_user_info(json.loads(token_json))
        service = build('gmail', 'v1', credentials=creds)

        # 查询最近7天的收件邮件
        results = service.users().messages().list(
            userId='me', q='in:inbox newer_than:7d', maxResults=100
        ).execute()

        messages = results.get('messages', [])
        replied_lead_ids = []

        for msg in messages:
            msg_detail = service.users().messages().get(
                userId='me', id=msg['id'], format='metadata',
                metadataHeaders=['From', 'Date']
            ).execute()

            headers = msg_detail.get('payload', {}).get('headers', [])
            from_header = next((h['value'] for h in headers if h['name'] == 'From'), '')
            date_header = next((h['value'] for h in headers if h['name'] == 'Date'), '')

            # 提取发件人邮箱
            import re
            match = re.search(r'<(.+?)>', from_header)
            sender = (match.group(1) if match else from_header).lower().strip()

            # 如果发件人在我们的leads列表中
            if sender in lead_emails:
                lead_id, last_sent_at = lead_emails[sender]

                # 获取邮件接收时间
                try:
                    from email.utils import parsedate_to_datetime
                    received_time = parsedate_to_datetime(date_header)

                    # 解析发送时间
                    from dateutil import parser as dateparser
                    sent_time = dateparser.parse(last_sent_at)

                    # 只有邮件是在我们发送之后收到的，才算作回复
                    if received_time > sent_time:
                        replied_lead_ids.append((lead_id, sender))
                        print(f"[Reply] {sender} replied (sent: {sent_time}, received: {received_time})")
                    else:
                        print(f"[Skip] {sender} email too old (sent: {sent_time}, received: {received_time})")
                except Exception as e:
                    print(f"[Date parse error] {e}, skipping message")
                    continue

        # 批量更新数据库（一次性提交，减少锁定时间）
        if replied_lead_ids:
            with get_db() as conn:
                for lead_id, sender in replied_lead_ids:
                    conn.execute("UPDATE leads SET replied=1, status='replied' WHERE id=?", (lead_id,))
                    print(f"[Reply detected] Lead {lead_id} ({sender}) replied")
                conn.commit()

    except Exception as e:
        print(f"[Check replies error] {e}")

def process_campaign(cid: int, account_email: str):
    # 先检查回复
    check_replies(cid, account_email)

    with get_db() as conn:
        lead = conn.execute("""
            SELECT l.* FROM leads l WHERE l.campaign_id=? AND l.status='pending' AND l.replied=0
            AND (l.last_sent_at IS NULL OR datetime(l.last_sent_at, '+' ||
                (SELECT delay_days FROM templates WHERE campaign_id=? AND step=l.current_step) || ' days') <= datetime('now'))
            LIMIT 1
        """, (cid, cid)).fetchone()
        if not lead: return

        tpl = conn.execute("SELECT * FROM templates WHERE campaign_id=? AND step=?",
                          (cid, lead['current_step'])).fetchone()
        if not tpl:
            conn.execute("UPDATE leads SET status='completed' WHERE id=?", (lead['id'],))
            conn.commit(); return

        data = json.loads(lead['data'])
        data['email'] = lead['email']
        subject = Template(tpl['subject']).render(**data)
        body = Template(tpl['body']).render(**data)

        # 添加追踪像素
        # 支持两种追踪服务：
        # 1. 使用Render上的tracker.py: TRACKER_URL=https://your-tracker.onrender.com
        # 2. 使用本地追踪服务: BASE_URL=http://your-vps-ip
        tracker_url = os.environ.get("TRACKER_URL", "")
        if tracker_url:
            # 使用外部追踪服务（如Render上的tracker.py）
            track_pixel = f'<img src="{tracker_url}/open?uid={lead["id"]}" width="1" height="1" style="display:none">'
        else:
            # 使用本地追踪服务
            base_url = os.environ.get("BASE_URL", "http://localhost:8000")
            track_pixel = f'<img src="{base_url}/track/open/{lead["id"]}" width="1" height="1" style="display:none">'
        body += track_pixel

        if send_gmail(account_email, lead['email'], subject, body):
            next_tpl = conn.execute("SELECT * FROM templates WHERE campaign_id=? AND step=?",
                                   (cid, lead['current_step'] + 1)).fetchone()
            new_status = 'pending' if next_tpl else 'completed'
            conn.execute("UPDATE leads SET current_step=current_step+1, last_sent_at=?, status=? WHERE id=?",
                        (datetime.now().isoformat(), new_status, lead['id']))
            conn.commit()

# 追踪
@app.get("/track/open/{lead_id}")
def track_open(lead_id: int, request: Request):
    """追踪邮件打开"""
    # 记录日志（包含IP和User-Agent）
    client_ip = request.client.host if request.client else "unknown"
    user_agent = request.headers.get("user-agent", "unknown")
    print(f"[Open tracked] Lead {lead_id} | IP: {client_ip} | UA: {user_agent[:50]}...")

    with get_db() as conn:
        # 检查lead是否存在
        lead = conn.execute("SELECT id, email FROM leads WHERE id=?", (lead_id,)).fetchone()
        if lead:
            conn.execute("UPDATE leads SET opened=1 WHERE id=?", (lead_id,))
            conn.commit()
            print(f"[Open tracked] ✓ Lead {lead_id} ({lead['email']}) marked as opened")
        else:
            print(f"[Open tracked] ✗ Lead {lead_id} not found")

    # 返回1x1透明GIF
    gif = base64.b64decode("R0lGODlhAQABAIAAAAAAAP///yH5BAEAAAAALAAAAAABAAEAAAIBRAA7")
    return HTMLResponse(content=gif, media_type="image/gif")

@app.get("/track/click/{lead_id}")
def track_click(lead_id: int, url: str):
    with get_db() as conn:
        conn.execute("UPDATE leads SET clicked=1 WHERE id=?", (lead_id,))
        conn.commit()
    return RedirectResponse(url)

@app.get("/api/accounts")
def list_accounts():
    with get_db() as conn:
        rows = conn.execute("SELECT id, email, created_at FROM accounts").fetchall()
        accounts = [dict(r) for r in rows]
        # 测试模式下添加虚拟账号
        if TEST_MODE and not accounts:
            conn.execute("INSERT OR IGNORE INTO accounts(email, token) VALUES(?, ?)",
                        ("test@example.com", "{}"))
            conn.commit()
            accounts = [{"id": 1, "email": "test@example.com", "created_at": "test"}]
        return accounts

@app.get("/api/campaigns/{cid}/stats")
def campaign_stats(cid: int):
    with get_db() as conn:
        stats = conn.execute("""
            SELECT COUNT(*) as total, SUM(opened) as opens, SUM(clicked) as clicks, SUM(replied) as replies,
            SUM(CASE WHEN status IN ('completed', 'replied') OR current_step > 1 THEN 1 ELSE 0 END) as sent
            FROM leads WHERE campaign_id=?
        """, (cid,)).fetchone()
        return dict(stats)

@app.post("/api/campaigns/{cid}/check-replies")
def check_replies_now(cid: int):
    """手动触发检查回复"""
    with get_db() as conn:
        campaign = conn.execute("SELECT account_email FROM campaigns WHERE id=?", (cid,)).fetchone()
        if not campaign or not campaign['account_email']:
            return {"msg": "请先启动campaign以设置发件账号"}
    check_replies(cid, campaign['account_email'])
    return {"msg": "检查完成"}

@app.post("/api/leads/{lead_id}/mark")
def mark_lead(lead_id: int, field: str):
    """手动标记lead状态（用于测试）"""
    if field not in ('opened', 'clicked', 'replied'):
        raise HTTPException(400, "Invalid field")
    with get_db() as conn:
        if field == 'replied':
            conn.execute("UPDATE leads SET replied=1, status='replied' WHERE id=?", (lead_id,))
        else:
            conn.execute(f"UPDATE leads SET {field}=1 WHERE id=?", (lead_id,))
        conn.commit()
    return {"ok": True}

# ============================================
# 自动同步追踪数据（从Render tracker拉取）
# ============================================

def sync_tracker_data():
    """
    从Render追踪服务同步数据到本地数据库
    每10分钟自动运行一次
    """
    import requests

    tracker_url = os.environ.get("TRACKER_URL", "")

    if not tracker_url or "your-tracker" in tracker_url:
        # 未配置追踪URL，跳过同步
        return

    try:
        # 获取未同步的打开记录
        response = requests.get(f"{tracker_url}/api/opens", timeout=10)
        if response.status_code != 200:
            print(f"[Sync Warning] Tracker API returned {response.status_code}")
            return

        data = response.json()
        opens = data.get("opens", [])

        if not opens:
            return

        # 更新本地数据库
        with get_db() as conn:
            updated = 0
            synced_ids = []

            for record in opens:
                uid = record['uid']
                record_id = record['id']

                try:
                    # 检查lead是否存在且未标记打开
                    lead = conn.execute(
                        "SELECT opened FROM leads WHERE id=?",
                        (uid,)
                    ).fetchone()

                    if lead and lead[0] == 0:
                        conn.execute("UPDATE leads SET opened=1 WHERE id=?", (uid,))
                        updated += 1

                    synced_ids.append(record_id)

                except Exception as e:
                    print(f"[Sync Error] Lead {uid}: {e}")

            conn.commit()

        # 标记远程记录为已同步
        if synced_ids:
            try:
                requests.post(
                    f"{tracker_url}/api/mark_synced",
                    json={"open_ids": synced_ids, "click_ids": []},
                    timeout=5
                )
                print(f"[Sync] Updated {updated} opens, marked {len(synced_ids)} as synced")
            except:
                pass  # 忽略标记失败

        # 同步点击记录
        response = requests.get(f"{tracker_url}/api/clicks", timeout=10)
        if response.status_code == 200:
            clicks = response.json().get("clicks", [])
            if clicks:
                with get_db() as conn:
                    click_ids = []
                    for record in clicks:
                        try:
                            conn.execute("UPDATE leads SET clicked=1 WHERE id=?", (record['uid'],))
                            click_ids.append(record['id'])
                        except:
                            pass
                    conn.commit()

                if click_ids:
                    requests.post(
                        f"{tracker_url}/api/mark_synced",
                        json={"open_ids": [], "click_ids": click_ids},
                        timeout=5
                    )

    except requests.exceptions.RequestException as e:
        print(f"[Sync Error] {e}")
    except Exception as e:
        print(f"[Sync Error] Unexpected: {e}")

# 添加自动同步任务（每10分钟）
if os.environ.get("TRACKER_URL"):
    scheduler.add_job(
        sync_tracker_data,
        'interval',
        minutes=10,
        id='sync_tracker',
        replace_existing=True
    )
    print("[Scheduler] Tracker sync enabled (every 10 minutes)")

# 前端页面
@app.get("/", response_class=HTMLResponse)
def index():
    return Path("index.html").read_text(encoding='utf-8')

# 启动时恢复运行中的 campaigns（延迟执行，确保所有函数已定义）
scheduler.add_job(restore_running_campaigns, 'date', id='restore_on_start')

if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
